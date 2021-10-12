import sys
from bs4 import BeautifulSoup
import shutil
import openpyxl
import glob
import re
import os

def insert_Excel(before_dir, after_dir, resultsFile, xlsxTemplate):
    shutil.copyfile(xlsxTemplate, resultsFile) # 結果を入れるエクセルを用意

    translatedList = [] # beforeのsrcとtargetが入る
    checkedList = [] # afterのsrcとtargetが入る

    before_dir = str(before_dir)
    after_dir = str(after_dir)

    before_html_files = glob.glob(before_dir + '/' + '*.html') # beforeのja-JP内にあるhtml(s)
    after_html_files = glob.glob(after_dir + '/' + '*.html') # afterのja-JP内にあるhtml(s)
    
    # xlfファイル名とbeforeのsrcとtargetを抽出
    for html in before_html_files:
        # xlfの名前を取得
        xlf = re.sub('\.html$', '', html)
        xlf_basename = os.path.basename(xlf)
        # print(xlf_basename)

        # 翻訳後のhtmlをオープンしてパース
        with open(html, encoding='utf-8') as f:
            translated = f.read()
        soupTranslated = BeautifulSoup(translated, 'html.parser')

        for t in soupTranslated.find_all('tr'):
            cols = t.find_all('td')
            src = re.sub(r'\t', ' ', cols[0].get_text()) # タブをスペースに置換
            target = re.sub(r'\t', ' ', cols[1].get_text()) # タブをスペースに置換
            translatedList.append(xlf_basename + '\t' + src + '\t' + target)
    
    # afterのsrcとtargetを抽出
    for html in after_html_files:
        # チェック後のhtmlをオープンしてパース
        with open(html, encoding='utf-8') as f:
            checked = f.read()
        soupChecked = BeautifulSoup(checked, 'html.parser')

        for t in soupChecked.find_all('tr'):
            cols = t.find_all('td')
            src = re.sub(r'\t', ' ', cols[0].get_text()) # タブをスペースに置換
            target = re.sub(r'\t', ' ', cols[1].get_text()) # タブをスペースに置換
            checkedList.append(src + '\t' + target)

    # Excelを準備
    wb = openpyxl.load_workbook(resultsFile)
    ws = wb['Sheet1']

    # 翻訳後のテキストを入力する
    countT = 2
    for i in translatedList:
        countStr = str(countT)
        f_name, src, target = i.split('\t')
        judge = '=IF(C'+countStr+'=F'+countStr+',"-","check!")'
        f_nameA = 'A' + countStr
        srcB = 'B' + countStr
        targetC = 'C' + countStr
        judgeD = 'D' + countStr
        ws[f_nameA].value = f_name
        ws[srcB].value = src
        ws[targetC].value = target
        ws[judgeD].value = judge
        countT += 1
    
    # チェック後のテキストを入力する
    countC = 2
    for i in checkedList:
        countStr = str(countC)
        src, target =i.split('\t')
        srcE = 'E' + countStr
        targetF = 'F' + countStr
        ws[srcE].value = src
        ws[targetF].value = target
        countC += 1

    # Excelを閉じて保存
    wb.close()
    wb.save(resultsFile)


if __name__ == '__main__':
    # 引数
    before_dir = sys.argv[1]
    after_dir = sys.argv[2]
    resultsFile = sys.argv[3]
    xlsxTemplate = sys.argv[4]

    # htmlの中身をExcelに入力
    insert_Excel(before_dir, after_dir, resultsFile, xlsxTemplate)
